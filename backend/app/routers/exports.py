from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from sqlalchemy.orm import selectinload
from typing import Optional
import csv
import io

from app.database import get_db
from app.models import StudentRecord

router = APIRouter()


@router.get("/csv")
async def export_csv(
    batch_id: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db)
):
    query = select(StudentRecord).options(selectinload(StudentRecord.evaluation))
    if batch_id:
        query = query.where(StudentRecord.batch_id == batch_id)

    result = await db.execute(query)
    records = result.scalars().all()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "Register Number", "Student Name", "Process Marks (18)",
        "Product Marks (12)", "Total (30)", "Remarks", "Status"
    ])

    for record in records:
        eval_data = record.evaluation
        writer.writerow([
            record.register_number or "N/A",
            record.student_name or "Unknown",
            eval_data.total_process if eval_data else 0,
            eval_data.total_product if eval_data else 0,
            eval_data.total_score if eval_data else 0,
            eval_data.remarks if eval_data else "",
            record.file_status,
        ])

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=practicum_results.csv"}
    )


@router.get("/xlsx")
async def export_xlsx(
    batch_id: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db)
):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        from fastapi import HTTPException
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    query = select(StudentRecord).options(selectinload(StudentRecord.evaluation))
    if batch_id:
        query = query.where(StudentRecord.batch_id == batch_id)

    result = await db.execute(query)
    records = result.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Practicum Results"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="6B21A8", end_color="6B21A8", fill_type="solid")
    headers = ["Register Number", "Student Name", "Process (18)", "Product (12)", "Total (30)", "Remarks", "Status"]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, record in enumerate(records, 2):
        eval_data = record.evaluation
        ws.cell(row=row_idx, column=1, value=record.register_number or "N/A")
        ws.cell(row=row_idx, column=2, value=record.student_name or "Unknown")
        ws.cell(row=row_idx, column=3, value=eval_data.total_process if eval_data else 0)
        ws.cell(row=row_idx, column=4, value=eval_data.total_product if eval_data else 0)
        ws.cell(row=row_idx, column=5, value=eval_data.total_score if eval_data else 0)
        ws.cell(row=row_idx, column=6, value=eval_data.remarks if eval_data else "")
        ws.cell(row=row_idx, column=7, value=record.file_status)

    # Auto-size columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col) + 2
        ws.column_dimensions[col[0].column_letter].width = max_len

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=practicum_results.xlsx"}
    )
